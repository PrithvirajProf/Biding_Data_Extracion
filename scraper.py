import os
import sys
import signal
import time
import logging
import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- Configuration ---
LOG_FILE = "scraping_log.txt"
EXCEL_FILE = "delaware_all_bids.xlsx"
BASE_URL = "https://mmp.delaware.gov/Bids/"

# --- Main ---

def main():
    """
    Main function to orchestrate the web scraping process.
    """
    setup_logging()
    driver = initialize_browser()
    setup_interrupt_handler(driver)

    try:
        processed_bid_ids = load_processed_bid_ids(EXCEL_FILE)
        navigate_to_bids_page(driver, BASE_URL)
        process_bid_categories(driver, processed_bid_ids)
    finally:
        driver.quit()
        logging.info("Scraping complete and browser closed.")

# --- Setup and Configuration ---

def setup_logging():
    """
    Configures the logging settings to output messages to both a file and the console.
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILE),
            logging.StreamHandler()
        ]
    )

def initialize_browser():
    """
    Initializes and returns a Selenium Chrome WebDriver with specified options.
    """
    options = Options()
    options.add_argument("--headless")  # Uncomment to run in the background
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def setup_interrupt_handler(driver):
    """
    Sets up a signal handler for graceful shutdown on Ctrl+C.
    """
    def handle_interrupt(signum, frame):
        logging.warning("Scraping interrupted by user (Ctrl+C). Shutting down gracefully.")
        if driver:
            driver.quit()
        sys.exit()
    signal.signal(signal.SIGINT, handle_interrupt)

# --- Excel File Handling ---

def load_processed_bid_ids(excel_file):
    """
    Loads a set of already processed Bid IDs from the specified Excel file
    to prevent duplicate scraping.
    """
    if not os.path.exists(excel_file):
        logging.info(f"No existing Excel file found. \nStarting fresh.")
        return set()

    try:
        df_existing = pd.read_excel(excel_file)
        if "Bid ID" in df_existing.columns:
            processed_ids = set(df_existing["Bid ID"].astype(str))
            logging.info(f"Loaded {len(processed_ids)} previously scraped bid IDs.")
            return processed_ids
        else:
            logging.warning("Column 'Bid ID' not found in Excel. Starting with an empty set.")
            return set()
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return set()

def append_to_excel(bid_data, excel_file):
    """
    Appends a new row of bid data to the Excel file. Creates the file with headers
    if it doesn't exist.
    """
    df_row = pd.DataFrame([bid_data])
    if not os.path.exists(excel_file):
        df_row.to_excel(excel_file, index=False)
        logging.info("Created new Excel file with headers.")
        return

    try:
        book = load_workbook(excel_file)
        sheet = book.active
        # Append without writing headers again
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_row.to_excel(writer, index=False, header=False, startrow=sheet.max_row)
    except PermissionError:
        logging.error(f"Permission denied: Could not write to '{excel_file}'. It may be open.")
    except Exception as e:
        logging.error(f"Failed to save data to Excel: {e}")

# --- Web Scraping Logic ---

def navigate_to_bids_page(driver, url):
    """
    Navigates the browser to the specified URL.
    """
    driver.get(url)
    logging.info(f"Navigated to {url}")

def process_bid_categories(driver, processed_bid_ids):
    """
    Iterates through specified bid categories and scrapes data from each.
    """
    wait = WebDriverWait(driver, 15)

    categories = [("Open","btnOpen"),("Recently Closed","btnClosed"),("Not Awarded","btnNotAwarded")]
    for category_name, tab_id in categories:
        logging.info(f"--- Processing category: {category_name} ---")
        try:
            wait.until(EC.element_to_be_clickable((By.ID, tab_id))).click()
            time.sleep(2)  # Allow time for the table to load
            scrape_bids_from_table(driver, wait, category_name, processed_bid_ids)
        except Exception as e:
            logging.error(f"Could not process category '{category_name}': {e}")

def scrape_bids_from_table(driver, wait, category_name, processed_bid_ids):
    """
    Scrapes all bid data from the currently active table, handling pagination.
    """
    while True:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#jqGridBids tbody tr")))
        rows = driver.find_elements(By.CSS_SELECTOR, "#jqGridBids tbody tr")

        for row in rows:
            try:
                process_bid_row(row, driver, wait, category_name, processed_bid_ids)
            except Exception as e:
                logging.error(f"Error processing a bid row: {e}")
                # Optionally, refresh or take other recovery actions
                continue

        if not navigate_to_next_page(driver):
            logging.info("No more pages to scrape in this category.")
            break

def process_bid_row(row, driver, wait, category_name, processed_bid_ids):
    """
    Extracts data from a single row, opens the details modal, scrapes modal data,
    and saves the combined information.
    """
    cells = row.find_elements(By.TAG_NAME, "td")
    if len(cells) < 7:
        return  # Skip malformed rows

    bid_id = cells[0].get_attribute("title")
    if str(bid_id) in processed_bid_ids:
        logging.info(f"Skipping already scraped Bid ID: {bid_id}")
        return

    # Extract data from the main table row
    bid_summary = {
        "Category": category_name,
        "Bid ID": bid_id,
        "Contract Number": cells[1].text.strip(),
        "Title": cells[2].text.strip(),
        "Open Date": cells[3].text.strip(),
        "Deadline": cells[4].text.strip(),
        "Agency": cells[5].text.strip(),
        "UNSPSC": cells[6].text.strip(),
    }

    # Open the details modal
    title_link = cells[2].find_element(By.TAG_NAME, "a")
    driver.execute_script("arguments[0].click();", title_link)
    
    # Scrape data from the modal
    modal = wait.until(EC.presence_of_element_located((By.ID, "dynamicDialogInnerHtml")))
    modal_data = extract_modal_data(modal)

    # Close the modal
    ActionChains(driver).send_keys(u'\ue00c').perform() # Simulates pressing the Escape key
    time.sleep(1)

    # Combine and save data
    full_bid_data = {**bid_summary, **modal_data}
    append_to_excel(full_bid_data, EXCEL_FILE)
    processed_bid_ids.add(str(bid_id))
    logging.info(f"Successfully scraped and saved Bid ID: {bid_id}")

def extract_modal_data(modal):
    """
    Extracts detailed information from the bid details modal.
    """
    # Helper function to safely find an element and get its text
    def get_text(xpath):
        try:
            return modal.find_element(By.XPATH, xpath).text.strip()
        except:
            return "N/A"

    # Extract document links
    documents = {}
    try:
        doc_links = modal.find_elements(By.XPATH, ".//div[@id='bidDocuments']//a")
        for link in doc_links:
            documents[link.text.strip()] = link.get_attribute("href")
    except Exception as e:
        logging.warning(f"Could not extract document links: {e}")

    return {
        "Contact Email": get_text(".//a[contains(@href, 'mailto')]"),
        "Solicitation Ad Date": get_text(".//label[preceding-sibling::label[contains(text(),'Solicitation Ad Date')]]"),
        "Deadline for Bid Responses": get_text(".//label[preceding-sibling::label[contains(text(),'Deadline for Bid Responses')]]"),
        "Important Message": get_text(".//h6[contains(@class, 'text-danger')]"),
        "Documents": str(documents) if documents else "N/A"
    }

def navigate_to_next_page(driver):
    """
    Clicks the 'next' button on the pagination control if it's available.
    Returns True if successful, False otherwise.
    """
    try:
        next_btn = driver.find_element(By.ID, "next_jqg1")
        if "disabled" in next_btn.get_attribute("class") or \
           "ui-jqgrid-disablePointerEvents" in next_btn.get_attribute("class"):
            return False
        
        driver.execute_script("arguments[0].click();", next_btn)
        time.sleep(2) # Wait for the next page to load
        return True
    except Exception as e:
        logging.error(f"Pagination failed or reached the end: {e}")
        return False

# --- Entry Point ---

if __name__ == "__main__":
    main()